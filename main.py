from __future__ import annotations

import calendar
import json
import sys
import traceback
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from tkinter import font as tkfont

from openpyxl import Workbook, load_workbook
try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None

APP_NAME = "订阅管理助手"
BASE_DIR = Path(__file__).resolve().parent
def _user_data_dir() -> Path:
    if sys.platform == "darwin":
        return Path.home() / "Library" / "Application Support" / APP_NAME / "data"
    if sys.platform.startswith("win"):
        appdata = Path.home() / "AppData" / "Roaming"
        return appdata / APP_NAME / "data"
    return Path.home() / f".{APP_NAME}" / "data"


DATA_DIR = _user_data_dir()
DATA_FILE = DATA_DIR / "subscriptions.json"
DEFAULT_IMPORT_XLSX = Path("/Users/guo/Documents/订阅管理系统_终极版.xlsx")

CYCLE_TO_MONTHS: Dict[str, int] = {"月": 1, "季": 3, "年": 12, "2年": 24}
DEFAULT_CATEGORIES = ["AI", "工具", "娱乐", "创作", "会员", "设备", "其他"]

# White minimal palette
BG_MAIN = "#ffffff"
BG_PANEL = "#ffffff"
BG_CARD = "#ffffff"
BG_SOFT = "#f8fafc"
BORDER = "#e5e7eb"
TEXT_MAIN = "#0f172a"
TEXT_SUB = "#64748b"
ACCENT = "#2563eb"
ACCENT_SOFT = "#eff6ff"
ACCENT_TEXT = "#1d4ed8"


def install_crash_handler() -> None:
    def _handler(exc_type, exc_value, exc_tb):
        try:
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            log_file = DATA_DIR / "crash.log"
            content = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with log_file.open("a", encoding="utf-8") as f:
                f.write(f"\n[{now}] Unhandled Exception\n")
                f.write(content)
                f.write("\n")
            try:
                root = tk.Tk()
                root.withdraw()
                messagebox.showerror(
                    "程序异常退出",
                    f"程序发生异常，日志已写入：\n{log_file}\n\n{exc_value}",
                )
                root.destroy()
            except Exception:
                pass
        except Exception:
            pass

    sys.excepthook = _handler


@dataclass
class Subscription:
    service: str
    category: str
    cycle: str
    price: float
    renew_date: Optional[date]
    frequent: str
    importance: str
    status: str
    remark: str
    icon_path: str = ""

    @property
    def months(self) -> int:
        return CYCLE_TO_MONTHS.get(self.cycle, 0)

    @property
    def monthly_cost(self) -> float:
        return self.price / self.months if self.months else 0.0

    @property
    def yearly_cost(self) -> float:
        return self.monthly_cost * 12

    @property
    def cancel_saving(self) -> float:
        return self.yearly_cost if self.status == "待取消" else 0.0

    @property
    def days_left(self) -> Optional[int]:
        if not self.renew_date:
            return None
        return (self.renew_date - date.today()).days

    @property
    def reminder(self) -> str:
        if not self.renew_date:
            return "未填写"
        d = self.days_left
        if d is None:
            return "未填写"
        if d < 0:
            return "已过期"
        if d <= 7:
            return "7天内续费"
        if d <= 30:
            return "30天内续费"
        return "正常"

    @property
    def reminder_level(self) -> str:
        if not self.renew_date:
            return "未填写"
        d = self.days_left
        if d is None:
            return "未填写"
        if d < 0:
            return "已过期"
        if d <= 7:
            return "紧急"
        if d <= 30:
            return "关注"
        return "正常"

    def to_dict(self) -> Dict:
        data = asdict(self)
        data["renew_date"] = self.renew_date.isoformat() if self.renew_date else None
        return data

    @staticmethod
    def from_dict(d: Dict) -> "Subscription":
        raw_date = d.get("renew_date")
        parsed: Optional[date] = None
        if raw_date:
            parsed = datetime.strptime(raw_date, "%Y-%m-%d").date()
        return Subscription(
            service=str(d.get("service", "")).strip(),
            category=str(d.get("category", "")).strip(),
            cycle=str(d.get("cycle", "月")).strip() or "月",
            price=float(d.get("price", 0) or 0),
            renew_date=parsed,
            frequent=str(d.get("frequent", "N")).strip() or "N",
            importance=str(d.get("importance", "中")).strip() or "中",
            status=str(d.get("status", "活跃")).strip() or "活跃",
            remark=str(d.get("remark", "")).strip(),
            icon_path=str(d.get("icon_path", "")).strip(),
        )


class DataStore:
    def __init__(self) -> None:
        self.subscriptions: List[Subscription] = []

    def load(self) -> None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        if DATA_FILE.exists():
            payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
            self.subscriptions = [Subscription.from_dict(x) for x in payload]
            return

        if DEFAULT_IMPORT_XLSX.exists():
            self.import_from_excel(DEFAULT_IMPORT_XLSX)
            self.save()
        else:
            self.subscriptions = []

    def save(self) -> None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        DATA_FILE.write_text(
            json.dumps([x.to_dict() for x in self.subscriptions], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def import_from_excel(self, path: Path) -> None:
        wb = load_workbook(path)
        ws = wb["订阅清单"]
        result: List[Subscription] = []
        for row in ws.iter_rows(min_row=2, max_col=16, values_only=True):
            service = row[0]
            if not service:
                continue
            if str(service).strip() == "合计":
                continue

            renew_date = self._parse_any_date(row[7])
            result.append(
                Subscription(
                    service=str(service).strip(),
                    category=str(row[1] or "").strip(),
                    cycle=str(row[2] or "月").strip() or "月",
                    price=float(row[3] or 0),
                    renew_date=renew_date,
                    frequent=str(row[8] or "N").strip() or "N",
                    importance=str(row[9] or "中").strip() or "中",
                    status=str(row[10] or "活跃").strip() or "活跃",
                    remark=str(row[14] or "").strip(),
                    icon_path=str(row[15] or "").strip(),
                )
            )
        self.subscriptions = result

    def export_to_excel(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "订阅清单"
        ws.append(
            [
                "服务",
                "类别",
                "计费周期",
                "价格(元)",
                "折算月数",
                "月均成本(自动)",
                "年成本(自动)",
                "下次续费日期",
                "是否常用",
                "重要性",
                "状态",
                "取消后年省(自动)",
                "续费提醒(自动)",
                "成本排名",
                "备注",
                "图标路径",
            ]
        )

        sorted_by_yearly = sorted(self.subscriptions, key=lambda x: x.yearly_cost, reverse=True)
        rank_map = {id(item): i + 1 for i, item in enumerate(sorted_by_yearly)}

        for s in self.subscriptions:
            ws.append(
                [
                    s.service,
                    s.category,
                    s.cycle,
                    round(s.price, 2),
                    s.months,
                    round(s.monthly_cost, 2),
                    round(s.yearly_cost, 2),
                    s.renew_date.isoformat() if s.renew_date else "",
                    s.frequent,
                    s.importance,
                    s.status,
                    round(s.cancel_saving, 2),
                    s.reminder,
                    rank_map[id(s)],
                    s.remark,
                    s.icon_path,
                ]
            )

        ws.append(
            [
                "合计",
                "",
                "",
                "",
                "",
                round(sum(x.monthly_cost for x in self.subscriptions), 2),
                round(sum(x.yearly_cost for x in self.subscriptions), 2),
                "",
                "",
                "",
                "",
                round(sum(x.cancel_saving for x in self.subscriptions), 2),
                "",
                "",
                "",
                "",
            ]
        )

        wb.save(path)

    @staticmethod
    def _parse_any_date(raw) -> Optional[date]:
        if raw is None or raw == "":
            return None
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        text = str(raw).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None


class SingleChoiceDialog(tk.Toplevel):
    def __init__(self, master, title: str, options: List[str], current: str, on_add=None):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.result: Optional[str] = None
        self.on_add = on_add

        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=12, pady=12)
        ttk.Label(frame, text="请选择一项").pack(anchor="w")

        self.listbox = tk.Listbox(frame, height=min(8, max(4, len(options))), exportselection=False)
        self.listbox.pack(fill="both", expand=True, pady=(6, 10))
        for idx, item in enumerate(options):
            self.listbox.insert("end", item)
            if item == current:
                self.listbox.selection_set(idx)
                self.listbox.see(idx)
        if not self.listbox.curselection() and options:
            self.listbox.selection_set(0)

        button_row = ttk.Frame(frame)
        button_row.pack(fill="x")
        if self.on_add is not None:
            ttk.Button(button_row, text="新增选项", command=self._add).pack(side="left")
        ttk.Button(button_row, text="取消", command=self.destroy).pack(side="left")
        ttk.Button(button_row, text="确定", command=self._confirm).pack(side="right")

        self.listbox.bind("<Double-Button-1>", lambda _e: self._confirm())
        self.listbox.bind("<Return>", lambda _e: self._confirm())

        self.transient(master)
        self.wait_visibility()
        self.focus_set()

    def _confirm(self) -> None:
        sel = self.listbox.curselection()
        if not sel:
            return
        self.result = str(self.listbox.get(sel[0]))
        self.destroy()

    def _add(self) -> None:
        if self.on_add is None:
            return
        new_value = self.on_add(self)
        if not new_value:
            return
        items = [self.listbox.get(i) for i in range(self.listbox.size())]
        if new_value not in items:
            self.listbox.insert("end", new_value)
        for i in range(self.listbox.size()):
            if self.listbox.get(i) == new_value:
                self.listbox.selection_clear(0, "end")
                self.listbox.selection_set(i)
                self.listbox.see(i)
                break


class DatePickerDialog(tk.Toplevel):
    def __init__(self, master, title: str, initial: Optional[date] = None):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.configure(bg=BG_MAIN)
        self.result: Optional[date] = None

        base = initial or date.today()
        self.year_var = tk.StringVar(value=f"{base.year}年")
        self.month_var = tk.StringVar(value=f"{base.month}月")
        self.day_var = tk.StringVar(value=f"{base.day}日")

        panel = ttk.LabelFrame(self, text="日期选择", style="Glass.TLabelframe")
        panel.pack(fill="both", expand=True, padx=14, pady=12)

        row = ttk.Frame(panel)
        row.pack(fill="x", padx=14, pady=(14, 10))

        ttk.Label(row, text="年", style="DialogLabel.TLabel").pack(side="left", padx=(0, 6))
        self.year_box = ttk.Combobox(
            row,
            textvariable=self.year_var,
            values=[f"{y}年" for y in range(base.year - 10, base.year + 11)],
            state="readonly",
            width=8,
        )
        self.year_box.pack(side="left", padx=(0, 18))

        ttk.Label(row, text="月", style="DialogLabel.TLabel").pack(side="left", padx=(0, 6))
        self.month_box = ttk.Combobox(
            row,
            textvariable=self.month_var,
            values=[f"{m}月" for m in range(1, 13)],
            state="readonly",
            width=5,
        )
        self.month_box.pack(side="left", padx=(0, 18))

        ttk.Label(row, text="日", style="DialogLabel.TLabel").pack(side="left", padx=(0, 6))
        self.day_box = ttk.Combobox(row, textvariable=self.day_var, values=[], state="readonly", width=5)
        self.day_box.pack(side="left")

        self.year_box.bind("<<ComboboxSelected>>", lambda _e: self._sync_day_options())
        self.month_box.bind("<<ComboboxSelected>>", lambda _e: self._sync_day_options())

        button_row = ttk.Frame(panel)
        button_row.pack(fill="x", padx=14, pady=(4, 12))
        ttk.Button(button_row, text="今天", width=10, command=self._set_today).pack(side="left")
        ttk.Button(button_row, text="取消", width=10, command=self.destroy).pack(side="right")
        ttk.Button(button_row, text="确定", width=10, command=self._confirm).pack(side="right", padx=(0, 8))

        self._sync_day_options()
        self.transient(master)
        self.wait_visibility()
        self.focus_set()

    def _set_today(self) -> None:
        today = date.today()
        self.year_var.set(f"{today.year}年")
        self.month_var.set(f"{today.month}月")
        self._sync_day_options()
        self.day_var.set(f"{today.day}日")

    def _sync_day_options(self) -> None:
        try:
            y = int(self.year_var.get().replace("年", ""))
            m = int(self.month_var.get().replace("月", ""))
            max_day = calendar.monthrange(y, m)[1]
            current_day = int(self.day_var.get().replace("日", "")) if self.day_var.get() else 1
            current_day = min(max(current_day, 1), max_day)
            self.day_box.configure(values=[f"{d}日" for d in range(1, max_day + 1)])
            self.day_var.set(f"{current_day}日")
        except Exception:
            return

    def _confirm(self) -> None:
        self._sync_day_options()
        try:
            y = int(self.year_var.get().replace("年", ""))
            m = int(self.month_var.get().replace("月", ""))
            d = int(self.day_var.get().replace("日", ""))
            self.result = date(y, m, d)
            self.destroy()
        except ValueError:
            messagebox.showerror("输入错误", "日期无效，请重新选择")


class SubscriptionDialog(tk.Toplevel):
    def __init__(self, master, title: str, initial: Optional[Subscription] = None):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.minsize(760, 640)
        self.configure(bg=BG_MAIN)
        self.result: Optional[Subscription] = None
        self.category_options = self._build_category_options(master, initial)
        self.cycle_options = self._build_cycle_options(master, initial)
        self.button_style = getattr(master, "base_button_style", "TButton")
        self.primary_button_style = getattr(master, "primary_button_style", "TButton")

        self.vars = {
            "service": tk.StringVar(value=initial.service if initial else ""),
            "category": tk.StringVar(
                value=(initial.category if initial and initial.category else self.category_options[0])
            ),
            "cycle": tk.StringVar(value=initial.cycle if initial else "月"),
            "price": tk.StringVar(value=str(initial.price) if initial else "0"),
            "renew_date": tk.StringVar(
                value=initial.renew_date.isoformat() if initial and initial.renew_date else ""
            ),
            "frequent": tk.StringVar(value=initial.frequent if initial else "Y"),
            "importance": tk.StringVar(value=initial.importance if initial else "中"),
            "status": tk.StringVar(value=initial.status if initial else "活跃"),
            "remark": tk.StringVar(value=initial.remark if initial else ""),
            "icon_path": tk.StringVar(value=initial.icon_path if initial else ""),
        }

        fields = [
            ("服务", "service"),
            ("类别", "category"),
            ("计费周期", "cycle"),
            ("价格(元)", "price"),
            ("下次续费日期(YYYY-MM-DD)", "renew_date"),
            ("是否常用", "frequent"),
            ("重要性", "importance"),
            ("状态", "status"),
            ("备注", "remark"),
            ("服务图标", "icon_path"),
        ]
        select_values = {
            "category": self.category_options,
            "cycle": self.cycle_options,
            "frequent": ["Y", "N"],
            "importance": ["高", "中", "低"],
            "status": ["活跃", "待取消"],
        }

        panel = ttk.LabelFrame(self, text="订阅信息", style="Glass.TLabelframe")
        panel.pack(fill="both", expand=True, padx=16, pady=(12, 8))

        for i, (label, key) in enumerate(fields):
            ttk.Label(panel, text=label, style="DialogLabel.TLabel").grid(row=i, column=0, sticky="w", padx=14, pady=8)
            if key in select_values:
                options = select_values[key]
                current = self.vars[key].get().strip()
                if current not in options and options:
                    self.vars[key].set(options[0])
                pick_frame = ttk.Frame(panel)
                pick_frame.grid(row=i, column=1, padx=10, pady=6, sticky="w")
                widget = ttk.Entry(pick_frame, textvariable=self.vars[key], width=34, state="readonly")
                widget.pack(side="left")
                ttk.Button(
                    pick_frame,
                    text="选择",
                    width=8,
                    style=self.button_style,
                    command=lambda k=key, opts=options: self._pick_option(k, opts),
                ).pack(side="left", padx=(6, 0))
            elif key == "renew_date":
                date_frame = ttk.Frame(panel)
                date_frame.grid(row=i, column=1, padx=10, pady=6, sticky="w")
                widget = ttk.Entry(date_frame, textvariable=self.vars[key], width=34, state="readonly")
                widget.pack(side="left")
                ttk.Button(date_frame, text="选择", width=8, style=self.button_style, command=self._pick_date).pack(side="left", padx=(6, 0))
                ttk.Button(date_frame, text="清空", width=8, style=self.button_style, command=lambda: self.vars["renew_date"].set("")).pack(
                    side="left", padx=(6, 0)
                )
            elif key == "icon_path":
                icon_frame = ttk.Frame(panel)
                icon_frame.grid(row=i, column=1, padx=10, pady=6, sticky="w")
                widget = ttk.Entry(icon_frame, textvariable=self.vars[key], width=34, state="readonly")
                widget.pack(side="left")
                ttk.Button(icon_frame, text="选择", width=8, style=self.button_style, command=self._choose_icon).pack(side="left", padx=(6, 0))
            else:
                widget = ttk.Entry(panel, textvariable=self.vars[key], width=46)
            if key not in select_values and key not in {"icon_path", "renew_date"}:
                widget.grid(row=i, column=1, padx=10, pady=6, sticky="w")

        tip = ttk.Label(self, text="提示：带“选择”按钮的字段为预设项，避免输入错误。", style="SubTitle.TLabel")
        tip.pack(anchor="w", padx=20, pady=(0, 8))

        button_frame = ttk.Frame(self)
        button_frame.pack(pady=(0, 14))
        ttk.Button(button_frame, text="取消", width=10, style=self.button_style, command=self.destroy).pack(side="left", padx=10)
        ttk.Button(button_frame, text="确定", width=10, style=self.primary_button_style, command=self._submit).pack(side="left", padx=10)

        self.transient(master)
        self.wait_visibility()
        self.focus_set()

    @staticmethod
    def _build_category_options(master, initial: Optional[Subscription]) -> List[str]:
        if hasattr(master, "dynamic_category_options") and master.dynamic_category_options:
            merged = list(master.dynamic_category_options)
            if initial and initial.category and initial.category not in merged:
                merged.append(initial.category)
            return merged
        existing = []
        if hasattr(master, "store") and hasattr(master.store, "subscriptions"):
            existing = [s.category for s in master.store.subscriptions if s.category]
        merged = []
        for x in DEFAULT_CATEGORIES + existing:
            if x and x not in merged:
                merged.append(x)
        if initial and initial.category and initial.category not in merged:
            merged.append(initial.category)
        return merged or DEFAULT_CATEGORIES

    @staticmethod
    def _build_cycle_options(master, initial: Optional[Subscription]) -> List[str]:
        if hasattr(master, "dynamic_cycle_options") and master.dynamic_cycle_options:
            options = list(master.dynamic_cycle_options)
        else:
            options = list(CYCLE_TO_MONTHS.keys())
        if initial and initial.cycle and initial.cycle not in options:
            options.append(initial.cycle)
        return options

    def _choose_icon(self) -> None:
        path = filedialog.askopenfilename(
            title="选择服务图标",
            filetypes=[
                ("图片文件", "*.png *.jpg *.jpeg *.gif *.bmp *.webp *.ppm *.pgm"),
                ("PNG/JPG", "*.png *.jpg *.jpeg"),
                ("所有文件", "*.*"),
            ],
        )
        if path:
            self.vars["icon_path"].set(path)

    def _pick_date(self) -> None:
        current = self.vars["renew_date"].get().strip()
        initial = None
        if current:
            try:
                initial = datetime.strptime(current, "%Y-%m-%d").date()
            except ValueError:
                initial = date.today()
        picker = DatePickerDialog(self, "选择日期", initial)
        self.wait_window(picker)
        if picker.result is not None:
            self.vars["renew_date"].set(picker.result.isoformat())

    def _pick_option(self, key: str, options: List[str]) -> None:
        names = {
            "category": "类别",
            "cycle": "计费周期",
            "frequent": "是否常用",
            "importance": "重要性",
            "status": "状态",
        }
        on_add = None
        if key == "category":
            on_add = self._add_category_option
        elif key == "cycle":
            on_add = self._add_cycle_option
        picker = SingleChoiceDialog(self, f"选择{names.get(key, key)}", options, self.vars[key].get().strip(), on_add=on_add)
        self.wait_window(picker)
        if picker.result:
            self.vars[key].set(picker.result)
            if key == "category" and picker.result not in self.category_options:
                self.category_options.append(picker.result)
            if key == "cycle" and picker.result not in self.cycle_options:
                self.cycle_options.append(picker.result)

    def _add_category_option(self, parent) -> Optional[str]:
        value = simpledialog.askstring("新增类别", "请输入新类别名称：", parent=parent)
        if not value:
            return None
        value = value.strip()
        if not value:
            return None
        if value not in self.category_options:
            self.category_options.append(value)
            if hasattr(self.master, "dynamic_category_options") and value not in self.master.dynamic_category_options:
                self.master.dynamic_category_options.append(value)
        return value

    def _add_cycle_option(self, parent) -> Optional[str]:
        name = simpledialog.askstring("新增计费周期", "请输入周期名称（例如：半年）", parent=parent)
        if not name:
            return None
        name = name.strip()
        if not name:
            return None
        months = simpledialog.askinteger("折算月数", f"请输入“{name}”对应的折算月数（整数）", parent=parent, minvalue=1, maxvalue=120)
        if not months:
            return None
        CYCLE_TO_MONTHS[name] = int(months)
        if name not in self.cycle_options:
            self.cycle_options.append(name)
        if hasattr(self.master, "dynamic_cycle_options") and name not in self.master.dynamic_cycle_options:
            self.master.dynamic_cycle_options.append(name)
        return name

    def _submit(self) -> None:
        try:
            price = float(self.vars["price"].get().strip() or "0")
        except ValueError:
            messagebox.showerror("输入错误", "价格必须是数字")
            return

        renew = self.vars["renew_date"].get().strip()
        parsed_date = None
        if renew:
            try:
                parsed_date = datetime.strptime(renew, "%Y-%m-%d").date()
            except ValueError:
                messagebox.showerror("输入错误", "日期格式必须是 YYYY-MM-DD")
                return

        service = self.vars["service"].get().strip()
        if not service:
            messagebox.showerror("输入错误", "服务名称不能为空")
            return

        cycle = self.vars["cycle"].get().strip()
        if cycle not in CYCLE_TO_MONTHS:
            messagebox.showerror("输入错误", "计费周期必须为：月 / 季 / 年 / 2年")
            return
        frequent = self.vars["frequent"].get().strip().upper()
        if frequent not in {"Y", "N"}:
            messagebox.showerror("输入错误", "是否常用必须为 Y 或 N")
            return
        importance = self.vars["importance"].get().strip()
        if importance not in {"高", "中", "低"}:
            messagebox.showerror("输入错误", "重要性必须为 高 / 中 / 低")
            return
        status = self.vars["status"].get().strip()
        if status not in {"活跃", "待取消"}:
            messagebox.showerror("输入错误", "状态必须为 活跃 / 待取消")
            return

        self.result = Subscription(
            service=service,
            category=self.vars["category"].get().strip(),
            cycle=cycle,
            price=price,
            renew_date=parsed_date,
            frequent=frequent,
            importance=importance,
            status=status,
            remark=self.vars["remark"].get().strip(),
            icon_path=self.vars["icon_path"].get().strip(),
        )
        self.destroy()


class RoundedValueBox(tk.Canvas):
    def __init__(self, master, text_var: tk.StringVar, height: int = 74, radius: int = 18, font=("Arial", 20, "bold")):
        super().__init__(master, height=height, highlightthickness=0, bd=0, bg=BG_MAIN)
        self.text_var = text_var
        self.radius = radius
        self.font = font
        self.text_var.trace_add("write", lambda *_: self._draw())
        self.bind("<Configure>", lambda _e: self._draw())
        self._draw()

    @staticmethod
    def _rounded_points(x1, y1, x2, y2, r):
        return [
            x1 + r,
            y1,
            x2 - r,
            y1,
            x2,
            y1,
            x2,
            y1 + r,
            x2,
            y2 - r,
            x2,
            y2,
            x2 - r,
            y2,
            x1 + r,
            y2,
            x1,
            y2,
            x1,
            y2 - r,
            x1,
            y1 + r,
            x1,
            y1,
        ]

    def _draw(self):
        self.delete("all")
        w = max(2, self.winfo_width())
        h = max(2, self.winfo_height())
        r = max(4, min(self.radius, w // 4, h // 2))
        points = self._rounded_points(2, 2, w - 2, h - 2, r)
        self.create_polygon(points, smooth=True, fill=BG_CARD, outline=BORDER, width=1)
        self.create_text(w / 2, h / 2, text=self.text_var.get(), fill=TEXT_MAIN, font=self.font)


class DrawTable(tk.Frame):
    def __init__(self, master, columns, headers, filterable_columns, on_header_click, show_inner_hscroll: bool = True):
        super().__init__(master, bg=BG_MAIN)
        self.columns = columns
        self.headers = headers
        self.filterable_columns = set(filterable_columns)
        self.on_header_click = on_header_click
        self.rows = []
        self.selected_row = None
        self.col_widths = {c: 80 for c in columns}
        self.header_h = 36
        self.row_h = 36
        self.images = {}
        self.external_xscroll = None

        self.header_canvas = tk.Canvas(self, height=self.header_h, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER, bd=0)
        self.body_canvas = tk.Canvas(self, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER, bd=0)
        self.v_scroll = ttk.Scrollbar(self, orient="vertical", command=self.body_canvas.yview)
        self.h_scroll = tk.Scrollbar(
            self,
            orient="horizontal",
            command=self._xview,
            bg="#e2e8f0",
            activebackground="#94a3b8",
            highlightthickness=0,
            bd=1,
            relief="sunken",
            width=16,
        )

        self.header_canvas.grid(row=0, column=0, sticky="ew")
        self.body_canvas.grid(row=1, column=0, sticky="nsew")
        self.v_scroll.grid(row=1, column=1, sticky="ns")
        if show_inner_hscroll:
            self.h_scroll.grid(row=2, column=0, sticky="ew")
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.body_canvas.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self._sync_xscroll)
        self.body_canvas.bind("<Button-1>", self._on_body_click)
        self.header_canvas.bind("<Button-1>", self._on_header_click)
        self.body_canvas.bind("<MouseWheel>", self._on_mousewheel, add="+")
        self.body_canvas.bind("<Button-4>", self._on_mousewheel, add="+")
        self.body_canvas.bind("<Button-5>", self._on_mousewheel, add="+")
        self.header_canvas.bind("<MouseWheel>", self._on_mousewheel, add="+")
        self.header_canvas.bind("<Button-4>", self._on_mousewheel, add="+")
        self.header_canvas.bind("<Button-5>", self._on_mousewheel, add="+")

    def set_external_xscroll(self, scrollbar) -> None:
        self.external_xscroll = scrollbar

    def _sync_xscroll(self, first, last):
        self.h_scroll.set(first, last)
        if self.external_xscroll is not None:
            self.external_xscroll.set(first, last)

    def _xview(self, *args):
        self.header_canvas.xview(*args)
        self.body_canvas.xview(*args)

    def _compute_widths(self):
        body_font = tkfont.nametofont("TkDefaultFont")
        head_font = tkfont.nametofont("TkHeadingFont")
        min_map = {"seq": 48, "icon": 56, "service": 110, "renew": 112, "reminder": 110, "remark": 90}
        max_w = 260
        pad = 20
        for col, head in zip(self.columns, self.headers):
            head_text = f"{head} ▾" if col in self.filterable_columns else head
            w = head_font.measure(head_text) + pad
            for row in self.rows:
                w = max(w, body_font.measure(str(row.get(col, ""))) + pad)
            w = max(min_map.get(col, 62), min(max_w, w))
            self.col_widths[col] = w

    def set_rows(self, rows, images):
        self.rows = rows
        self.images = images or {}
        self._compute_widths()
        self.selected_row = None if not rows else min(self.selected_row or 0, len(rows) - 1)
        self._redraw()

    def _total_width(self):
        return sum(self.col_widths[c] for c in self.columns)

    def _redraw(self):
        self.header_canvas.delete("all")
        self.body_canvas.delete("all")

        x = 0
        for col, head in zip(self.columns, self.headers):
            w = self.col_widths[col]
            self.header_canvas.create_rectangle(x, 0, x + w, self.header_h, outline=BORDER, fill=ACCENT_SOFT)
            text = f"{head} ▾" if col in self.filterable_columns else head
            self.header_canvas.create_text(x + w / 2, self.header_h / 2, text=text, fill=ACCENT_TEXT, font=("Arial", 12, "bold"))
            x += w
        self.header_canvas.configure(scrollregion=(0, 0, x, self.header_h))

        for r, row in enumerate(self.rows):
            y0 = r * self.row_h
            y1 = y0 + self.row_h
            bg = "#dbeafe" if self.selected_row == r else BG_CARD
            self.body_canvas.create_rectangle(0, y0, x, y1, outline=BORDER, fill=bg)
            cx = 0
            for col in self.columns:
                w = self.col_widths[col]
                self.body_canvas.create_line(cx + w, y0, cx + w, y1, fill=BORDER)
                if col == "icon":
                    icon = row.get("__icon")
                    if icon is not None:
                        self.body_canvas.create_image(cx + w / 2, y0 + self.row_h / 2, image=icon)
                else:
                    txt = str(row.get(col, ""))
                    color = "#1d4ed8" if col == "service" else TEXT_MAIN
                    font = ("Arial", 12, "bold") if col == "service" else ("Arial", 12)
                    self.body_canvas.create_text(cx + w / 2, y0 + self.row_h / 2, text=txt, fill=color, font=font)
                cx += w
        self.body_canvas.configure(scrollregion=(0, 0, x, len(self.rows) * self.row_h))

    def _column_at_x(self, x):
        cur = 0
        for col in self.columns:
            w = self.col_widths[col]
            if cur <= x < cur + w:
                return col
            cur += w
        return None

    def _on_header_click(self, event):
        x = self.header_canvas.canvasx(event.x)
        col = self._column_at_x(x)
        if col and col in self.filterable_columns:
            self.on_header_click(col, int(event.x))

    def _on_body_click(self, event):
        y = self.body_canvas.canvasy(event.y)
        row = int(y // self.row_h)
        if 0 <= row < len(self.rows):
            self.selected_row = row
            self._redraw()

    def _on_mousewheel(self, event):
        if getattr(event, "num", None) == 4:
            step = -2
        elif getattr(event, "num", None) == 5:
            step = 2
        else:
            delta = getattr(event, "delta", 0)
            if sys.platform == "darwin":
                step = int(-delta / 4) if delta else 0
            else:
                step = int(-delta / 120) if delta else 0
            if step == 0 and delta:
                step = -1 if delta > 0 else 1
        self.body_canvas.yview_scroll(step, "units")
        return "break"

    def get_selected_source_index(self):
        if self.selected_row is None or not (0 <= self.selected_row < len(self.rows)):
            return None
        return self.rows[self.selected_row].get("__index")


class App(tk.Tk):
    def __init__(self, store: DataStore):
        super().__init__()
        self.store = store
        self.icon_cache: Dict[str, tk.PhotoImage] = {}
        self.undo_stack: List[List[Dict]] = []
        self.dynamic_category_options = list(DEFAULT_CATEGORIES)
        self.dynamic_cycle_options = list(CYCLE_TO_MONTHS.keys())
        for s in self.store.subscriptions:
            if s.category and s.category not in self.dynamic_category_options:
                self.dynamic_category_options.append(s.category)
            if s.cycle and s.cycle not in self.dynamic_cycle_options:
                self.dynamic_cycle_options.append(s.cycle)
        self.base_button_style = "Glass.TButton"
        self.primary_button_style = "Accent.TButton"
        today = date.today()
        self.calendar_year = today.year
        self.calendar_month = today.month
        self.header_filter_close_bindid = None
        self.calendar_tooltip = None
        self.calendar_tooltip_after = None
        self.calendar_cell_meta: Dict[tuple, Dict] = {}
        self.title(APP_NAME)
        self.geometry("1180x760")
        self.minsize(1050, 680)
        self.configure(bg=BG_MAIN)

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        self._setup_styles(style)

        self._build_ui()
        self.refresh_all()

    def _setup_styles(self, style: ttk.Style) -> None:
        families = set(tkfont.families())
        primary = "SF Pro Text" if "SF Pro Text" in families else ("PingFang SC" if "PingFang SC" in families else "Arial")

        style.configure(".", background=BG_MAIN, foreground=TEXT_MAIN, font=(primary, 12))
        style.configure("TFrame", background=BG_MAIN)
        style.configure("TLabel", background=BG_MAIN, foreground=TEXT_MAIN, font=(primary, 12))
        style.configure("Title.TLabel", background=BG_MAIN, foreground=TEXT_MAIN, font=(primary, 20, "bold"))
        style.configure("SubTitle.TLabel", background=BG_MAIN, foreground=TEXT_SUB, font=(primary, 12))
        style.configure("DialogLabel.TLabel", background=BG_PANEL, foreground=TEXT_SUB, font=(primary, 12))

        style.configure("TNotebook", background=BG_MAIN, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(14, 8), background=BG_SOFT, foreground=TEXT_SUB, font=(primary, 12, "bold"))
        style.map(
            "TNotebook.Tab",
            background=[("selected", BG_CARD)],
            foreground=[("selected", TEXT_MAIN)],
        )

        style.configure("Glass.TLabelframe", background=BG_PANEL, bordercolor=BORDER, borderwidth=1, relief="solid")
        style.configure("Glass.TLabelframe.Label", background=BG_PANEL, foreground=TEXT_MAIN, font=(primary, 12, "bold"))

        style.configure("Card.TLabelframe", background=BG_CARD, bordercolor=BORDER, borderwidth=1, relief="solid")
        style.configure("Card.TLabelframe.Label", background=BG_CARD, foreground=TEXT_SUB, font=(primary, 12, "bold"))

        style.configure(
            "Glass.TButton",
            background=BG_CARD,
            foreground=TEXT_MAIN,
            bordercolor=BORDER,
            borderwidth=1,
            relief="raised",
            padding=(14, 8),
            font=(primary, 12, "bold"),
        )
        style.map("Glass.TButton", background=[("active", "#f3f4f6"), ("pressed", "#e5e7eb")])
        style.configure(
            "Accent.TButton",
            background=ACCENT,
            foreground="#ffffff",
            bordercolor=ACCENT,
            borderwidth=1,
            relief="raised",
            padding=(14, 8),
            font=(primary, 12, "bold"),
        )
        style.map("Accent.TButton", background=[("active", "#1d4ed8"), ("pressed", "#1e40af")])
        style.configure(
            "TCombobox",
            fieldbackground=BG_CARD,
            background=BG_CARD,
            foreground=TEXT_MAIN,
            bordercolor=BORDER,
            lightcolor=BORDER,
            darkcolor=BORDER,
            arrowcolor=TEXT_SUB,
        )
        style.configure("TEntry", fieldbackground=BG_CARD, foreground=TEXT_MAIN)

        style.configure(
            "Glass.Treeview",
            background=BG_CARD,
            fieldbackground=BG_CARD,
            foreground=TEXT_MAIN,
            rowheight=30,
            borderwidth=1,
            relief="solid",
            font=(primary, 12),
        )
        style.configure("Glass.Treeview.Heading", background=ACCENT_SOFT, foreground=ACCENT_TEXT, font=(primary, 12, "bold"))
        style.map("Glass.Treeview", background=[("selected", "#dce9ff")], foreground=[("selected", TEXT_MAIN)])

        style.configure("Dashboard.Treeview", font=(primary, 12), rowheight=30, background=BG_CARD, fieldbackground=BG_CARD)
        style.configure("Dashboard.Treeview.Heading", font=(primary, 12, "bold"), background=ACCENT_SOFT, foreground=ACCENT_TEXT)

    def _build_ui(self) -> None:
        top_bar = ttk.Frame(self)
        top_bar.pack(fill="x", padx=12, pady=10)
        ttk.Label(top_bar, text=APP_NAME, style="Title.TLabel").pack(side="left")

        buttons = ttk.Frame(top_bar)
        buttons.pack(side="right")
        ttk.Button(buttons, text="下载模板", style=self.base_button_style, command=self.download_template).pack(side="left", padx=4)
        ttk.Button(buttons, text="导入Excel", style=self.base_button_style, command=self.import_excel).pack(side="left", padx=4)
        ttk.Button(buttons, text="导出Excel", style=self.base_button_style, command=self.export_excel).pack(side="left", padx=4)
        ttk.Button(buttons, text="保存", style=self.primary_button_style, command=self.save).pack(side="left", padx=4)

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.notebook = notebook

        self.list_frame = ttk.Frame(notebook)
        self.dashboard_frame = ttk.Frame(notebook)
        self.reminder_frame = ttk.Frame(notebook)

        notebook.add(self.list_frame, text="订阅清单")
        notebook.add(self.dashboard_frame, text="仪表盘")
        notebook.add(self.reminder_frame, text="续费提醒")

        self._build_list_tab()
        self._build_dashboard_tab()
        self._build_reminder_tab()
        self._bind_mousewheel_scroll()

    def _bind_mousewheel_scroll(self) -> None:
        self.bind_all("<MouseWheel>", self._on_global_mousewheel, add="+")
        self.bind_all("<Button-4>", self._on_global_mousewheel, add="+")
        self.bind_all("<Button-5>", self._on_global_mousewheel, add="+")

    def _on_global_mousewheel(self, event) -> None:
        if not hasattr(self, "dashboard_canvas") or not hasattr(self, "notebook"):
            return
        if self.notebook.select() != str(self.dashboard_frame):
            return
        target = self.winfo_containing(event.x_root, event.y_root)
        if target is None:
            return
        if not str(target).startswith(str(self.dashboard_frame)):
            return

        if getattr(event, "num", None) == 4:
            delta_units = -2
        elif getattr(event, "num", None) == 5:
            delta_units = 2
        else:
            if sys.platform == "darwin":
                # macOS touchpad/scroll-wheel delta is already fine-grained
                delta_units = int(-event.delta / 3) if event.delta else 0
            else:
                delta_units = int(-event.delta / 120) if event.delta else 0
            if delta_units == 0 and event.delta:
                delta_units = -1 if event.delta > 0 else 1
        self.dashboard_canvas.yview_scroll(delta_units, "units")

    def _build_list_tab(self) -> None:
        ops = ttk.Frame(self.list_frame)
        ops.pack(fill="x", padx=8, pady=8)
        ttk.Button(ops, text="新增", style=self.base_button_style, command=self.add_item).pack(side="left", padx=4)
        ttk.Button(ops, text="编辑", style=self.base_button_style, command=self.edit_item).pack(side="left", padx=4)
        ttk.Button(ops, text="删除", style=self.base_button_style, command=self.delete_item).pack(side="left", padx=4)
        self.undo_btn = ttk.Button(ops, text="撤回", style=self.base_button_style, command=self.undo_last, state="disabled")
        self.undo_btn.pack(side="left", padx=4)

        filters = ttk.Frame(self.list_frame)
        filters.pack(fill="x", padx=8, pady=(0, 8))
        self.keyword_var = tk.StringVar()
        self.column_filters: Dict[str, str] = {}

        ttk.Label(filters, text="关键词").pack(side="left")
        keyword_entry = ttk.Entry(filters, textvariable=self.keyword_var, width=28)
        keyword_entry.pack(side="left", padx=(6, 10))
        ttk.Button(filters, text="重置", style=self.base_button_style, command=self.reset_filters).pack(side="left")
        self.filter_hint_var = tk.StringVar(value="提示：点击表头可筛选（类别/周期/常用/重要性/状态/续费提醒）")
        ttk.Label(filters, textvariable=self.filter_hint_var, style="SubTitle.TLabel").pack(side="left", padx=(14, 0))

        keyword_entry.bind("<KeyRelease>", lambda _e: self.refresh_list())

        columns = (
            "seq",
            "icon",
            "service",
            "category",
            "cycle",
            "price",
            "monthly",
            "yearly",
            "renew",
            "frequent",
            "importance",
            "status",
            "saving",
            "reminder",
            "remark",
        )
        headers = [
            "序号",
            "图标",
            "服务",
            "类别",
            "周期",
            "价格",
            "月均",
            "年成本",
            "续费日期",
            "常用",
            "重要性",
            "状态",
            "取消后年省",
            "续费提醒",
            "备注",
        ]
        self.filterable_columns = {"category", "cycle", "frequent", "importance", "status", "reminder"}

        bottom_area = ttk.Frame(self.list_frame)
        bottom_area.pack(side="bottom", fill="x", padx=8, pady=(0, 8))

        table_wrap = ttk.Frame(self.list_frame)
        table_wrap.pack(fill="both", expand=True, padx=8, pady=(0, 0))
        self.list_table_wrap = table_wrap

        self.draw_table = DrawTable(
            table_wrap,
            columns=columns,
            headers=headers,
            filterable_columns=self.filterable_columns,
            on_header_click=lambda col, x: self._show_header_dropdown(col, x),
            show_inner_hscroll=False,
        )
        self.draw_table.pack(fill="both", expand=True)
        self.list_x_scroll = tk.Scrollbar(
            bottom_area,
            orient="horizontal",
            command=self.draw_table._xview,
            bg="#e2e8f0",
            activebackground="#94a3b8",
            highlightthickness=0,
            bd=1,
            relief="sunken",
            width=16,
        )
        self.draw_table.set_external_xscroll(self.list_x_scroll)

        summary = ttk.LabelFrame(bottom_area, text="筛选结果合计", style="Glass.TLabelframe")
        summary.pack(fill="x")
        self.list_summary_vars = {
            "count": tk.StringVar(value="0"),
            "month": tk.StringVar(value="0.00"),
            "year": tk.StringVar(value="0.00"),
            "saving": tk.StringVar(value="0.00"),
        }
        ttk.Label(summary, text="条数:").pack(side="left", padx=(10, 4), pady=8)
        ttk.Label(summary, textvariable=self.list_summary_vars["count"]).pack(side="left", padx=(0, 14))
        ttk.Label(summary, text="月均合计:").pack(side="left", padx=(0, 4))
        ttk.Label(summary, textvariable=self.list_summary_vars["month"]).pack(side="left", padx=(0, 14))
        ttk.Label(summary, text="年成本合计:").pack(side="left", padx=(0, 4))
        ttk.Label(summary, textvariable=self.list_summary_vars["year"]).pack(side="left", padx=(0, 14))
        ttk.Label(summary, text="待取消可省/年:").pack(side="left", padx=(0, 4))
        ttk.Label(summary, textvariable=self.list_summary_vars["saving"]).pack(side="left")
        self.list_x_scroll.pack(fill="x", pady=(6, 0), side="bottom")

    def _build_dashboard_tab(self) -> None:
        outer = ttk.Frame(self.dashboard_frame)
        outer.pack(fill="both", expand=True)

        self.dashboard_canvas = tk.Canvas(outer, bg=BG_MAIN, highlightthickness=0, bd=0)
        y_scroll = ttk.Scrollbar(outer, orient="vertical", command=self.dashboard_canvas.yview)
        self.dashboard_canvas.configure(yscrollcommand=y_scroll.set)
        self.dashboard_canvas.pack(side="left", fill="both", expand=True)
        y_scroll.pack(side="right", fill="y")

        self.dashboard_content = ttk.Frame(self.dashboard_canvas)
        self.dashboard_window = self.dashboard_canvas.create_window((0, 0), window=self.dashboard_content, anchor="nw")

        def _on_content_configure(_event):
            self.dashboard_canvas.configure(scrollregion=self.dashboard_canvas.bbox("all"))

        def _on_canvas_configure(event):
            self.dashboard_canvas.itemconfigure(self.dashboard_window, width=event.width)

        self.dashboard_content.bind("<Configure>", _on_content_configure)
        self.dashboard_canvas.bind("<Configure>", _on_canvas_configure)

        calendar_box = ttk.LabelFrame(self.dashboard_content, text="续费日历视图", style="Glass.TLabelframe")
        calendar_box.pack(fill="x", padx=10, pady=(10, 10))

        nav = ttk.Frame(calendar_box)
        nav.pack(fill="x", padx=8, pady=(8, 4))
        self.calendar_summary_var = tk.StringVar(value="")
        tk.Label(nav, text="时间", bg=BG_PANEL, fg=TEXT_SUB, font=("Arial", 12, "bold")).pack(side="left", padx=(4, 8))
        nav_select = ttk.Frame(nav)
        nav_select.pack(side="left")
        self.calendar_year_pick_var = tk.StringVar()
        self.calendar_year_pick = ttk.Combobox(
            nav_select,
            textvariable=self.calendar_year_pick_var,
            values=[],
            state="readonly",
            width=6,
            font=("Arial", 16, "bold"),
        )
        self.calendar_year_pick.pack(side="left", padx=(0, 8))
        self.calendar_year_pick.bind("<<ComboboxSelected>>", self._on_calendar_year_pick)
        self.calendar_month_pick_var = tk.StringVar()
        self.calendar_month_pick = ttk.Combobox(
            nav_select,
            textvariable=self.calendar_month_pick_var,
            values=[f"{i:02d}" for i in range(1, 13)],
            state="readonly",
            width=4,
            font=("Arial", 16, "bold"),
        )
        self.calendar_month_pick.pack(side="left", padx=(0, 6))
        self.calendar_month_pick.bind("<<ComboboxSelected>>", self._on_calendar_month_pick)
        ttk.Button(nav_select, text="上月", width=8, style=self.base_button_style, command=lambda: self._change_calendar_month(-1)).pack(
            side="left", padx=(8, 4)
        )
        ttk.Button(nav_select, text="下月", width=8, style=self.base_button_style, command=lambda: self._change_calendar_month(1)).pack(
            side="left", padx=(0, 10)
        )
        ttk.Label(nav, textvariable=self.calendar_summary_var, font=("Arial", 13, "bold"), background=BG_PANEL, foreground=TEXT_SUB).pack(side="right")

        self.calendar_grid = tk.Frame(calendar_box, bg=BORDER)
        self.calendar_grid.pack(fill="x", padx=8, pady=(0, 8))
        for col in range(7):
            self.calendar_grid.grid_columnconfigure(col, weight=1, uniform="calendar")

        weekdays = ["一", "二", "三", "四", "五", "六", "日"]
        for col, name in enumerate(weekdays):
            head = tk.Label(
                self.calendar_grid,
                text=name,
                bg="#dbeafe",
                fg="#1e3a8a",
                relief="solid",
                bd=1,
                pady=6,
                font=("Arial", 12, "bold"),
            )
            head.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

        self.calendar_cells = []
        for r in range(1, 7):
            self.calendar_grid.grid_rowconfigure(r, weight=0, minsize=86)
            row_cells = []
            for c in range(7):
                cell = tk.Frame(self.calendar_grid, bg=BG_CARD, bd=1, relief="solid")
                cell.grid(row=r, column=c, sticky="nsew", padx=1, pady=1)
                day_label = tk.Label(cell, text="", anchor="nw", bg=BG_CARD, fg=TEXT_SUB, font=("Arial", 12, "bold"))
                day_label.pack(anchor="nw", padx=5, pady=(3, 0))
                icon_label = tk.Label(cell, text="", bg=BG_CARD, bd=0)
                icon_label.place(relx=1.0, x=-6, y=6, anchor="ne")
                icon_label2 = tk.Label(cell, text="", bg=BG_CARD, bd=0)
                icon_label2.place(relx=1.0, x=-28, y=6, anchor="ne")
                icon_more_label = tk.Label(cell, text="", bg=BG_CARD, fg=TEXT_SUB, font=("Arial", 9, "bold"))
                icon_more_label.place(relx=1.0, x=-52, y=8, anchor="ne")
                info_label = tk.Label(
                    cell,
                    text="",
                    anchor="nw",
                    justify="left",
                    bg=BG_CARD,
                    fg=TEXT_MAIN,
                    font=("Arial", 12, "bold"),
                    wraplength=140,
                )
                info_label.pack(anchor="nw", padx=5, pady=(1, 0))
                amount_label = tk.Label(
                    cell,
                    text="",
                    anchor="w",
                    justify="left",
                    bg=BG_CARD,
                    fg="#166534",
                    font=("Arial", 12, "bold"),
                )
                amount_label.pack(anchor="w", padx=5, pady=(0, 2))
                for widget in (cell, day_label, icon_label, icon_label2, icon_more_label, info_label, amount_label):
                    widget.bind("<Enter>", lambda e, rr=r - 1, cc=c: self._on_calendar_cell_enter(e, rr, cc), add="+")
                    widget.bind("<Leave>", lambda e: self._on_calendar_cell_leave(e), add="+")
                    widget.bind("<Motion>", lambda e, rr=r - 1, cc=c: self._on_calendar_cell_motion(e, rr, cc), add="+")
                row_cells.append((cell, day_label, icon_label, icon_label2, icon_more_label, info_label, amount_label))
            self.calendar_cells.append(row_cells)

        summary = ttk.Frame(self.dashboard_content)
        summary.pack(fill="x", padx=10, pady=(0, 8))

        self.metrics: Dict[str, tk.StringVar] = {
            "month": tk.StringVar(),
            "year": tk.StringVar(),
            "saving": tk.StringVar(),
            "active": tk.StringVar(),
            "pending": tk.StringVar(),
            "low_use": tk.StringVar(),
        }

        cards = [
            ("总月订阅", "month"),
            ("总年订阅", "year"),
            ("待取消可省/年", "saving"),
            ("活跃订阅数", "active"),
            ("待取消数量", "pending"),
            ("低使用率数量", "low_use"),
        ]

        for i, (title, key) in enumerate(cards):
            frame = ttk.Frame(summary)
            frame.grid(row=i // 3, column=i % 3, padx=8, pady=8, sticky="nsew")
            tk.Label(frame, text=title, bg=BG_MAIN, fg=TEXT_SUB, font=("Arial", 12, "bold")).pack(anchor="w")
            box = RoundedValueBox(frame, self.metrics[key], height=76, radius=18, font=("Arial", 21, "bold"))
            box.pack(fill="x", pady=(6, 0))

        summary.columnconfigure((0, 1, 2), weight=1)

        body = ttk.Frame(self.dashboard_content)
        body.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        left = ttk.LabelFrame(body, text="分类统计", style="Glass.TLabelframe")
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        right = ttk.LabelFrame(body, text="高成本 TOP5", style="Glass.TLabelframe")
        right.pack(side="left", fill="both", expand=True, padx=(8, 0))

        self.category_tree = ttk.Treeview(
            left,
            columns=("category", "monthly", "yearly"),
            show="headings",
            height=8,
            style="Dashboard.Treeview",
        )
        for col, title, w in (("category", "类别", 120), ("monthly", "月均成本", 120), ("yearly", "年成本", 120)):
            self.category_tree.heading(col, text=title)
            self.category_tree.column(col, width=w, anchor="center")
        self.category_tree.pack(fill="both", expand=True, padx=8, pady=8)

        self.top_tree = ttk.Treeview(
            right,
            columns=("rank", "service", "yearly"),
            show="headings",
            height=8,
            style="Dashboard.Treeview",
        )
        for col, title, w in (("rank", "排名", 80), ("service", "服务", 180), ("yearly", "年成本", 120)):
            self.top_tree.heading(col, text=title)
            self.top_tree.column(col, width=w, anchor="center")
        self.top_tree.pack(fill="both", expand=True, padx=8, pady=8)

    @staticmethod
    def _format_calendar_event_text(events: List[Subscription]) -> str:
        if not events:
            return ""
        names = [s.service for s in events if s.service]
        if len(names) <= 2:
            return "\n".join(names)
        return f"{names[0]}\n{names[1]}  +{len(names) - 2}项"

    def _change_calendar_month(self, delta: int) -> None:
        month = self.calendar_month + delta
        year = self.calendar_year
        if month <= 0:
            month = 12
            year -= 1
        elif month > 12:
            month = 1
            year += 1
        self.calendar_month = month
        self.calendar_year = year
        self._refresh_dashboard_calendar()

    def _on_calendar_month_pick(self, _event=None) -> None:
        raw = self.calendar_month_pick_var.get().strip().replace("月", "")
        try:
            picked = int(raw)
        except ValueError:
            return
        if 1 <= picked <= 12:
            self.calendar_month = picked
            self._refresh_dashboard_calendar()

    def _on_calendar_year_pick(self, _event=None) -> None:
        raw = self.calendar_year_pick_var.get().strip().replace("年", "")
        try:
            picked = int(raw)
        except ValueError:
            return
        self.calendar_year = picked
        self._refresh_dashboard_calendar()

    def _refresh_dashboard_calendar(self) -> None:
        subs = [s for s in self.store.subscriptions if s.renew_date is not None]
        month_subs = [s for s in subs if s.renew_date.year == self.calendar_year and s.renew_date.month == self.calendar_month]
        month_total = sum(s.price for s in month_subs)
        self.calendar_summary_var.set(f"本月续费 {len(month_subs)} 项  ¥{month_total:.2f}")
        years = sorted({s.renew_date.year for s in subs} | {self.calendar_year})
        if years:
            lo = min(years[0], self.calendar_year - 2)
            hi = max(years[-1], self.calendar_year + 2)
            years = list(range(lo, hi + 1))
        else:
            years = [self.calendar_year]
        if hasattr(self, "calendar_year_pick"):
            self.calendar_year_pick.configure(values=[str(y) for y in years])
        if hasattr(self, "calendar_year_pick_var"):
            self.calendar_year_pick_var.set(str(self.calendar_year))
        if hasattr(self, "calendar_month_pick_var"):
            self.calendar_month_pick_var.set(f"{self.calendar_month:02d}")

        day_map: Dict[int, List[Subscription]] = {}
        for s in month_subs:
            day_map.setdefault(s.renew_date.day, []).append(s)
        for x in day_map.values():
            x.sort(key=lambda i: i.service)
        self.calendar_cell_meta.clear()

        month_rows = calendar.monthcalendar(self.calendar_year, self.calendar_month)
        while len(month_rows) < 6:
            month_rows.append([0] * 7)

        for r in range(6):
            for c in range(7):
                day_num = month_rows[r][c]
                cell, day_label, icon_label, icon_label2, icon_more_label, info_label, amount_label = self.calendar_cells[r][c]
                if day_num == 0:
                    cell.configure(bg=BG_SOFT)
                    day_label.configure(text="", bg=BG_SOFT, fg="#a3b2d6")
                    icon_label.configure(image="", text="", bg=BG_SOFT)
                    icon_label.image = None
                    icon_label2.configure(image="", text="", bg=BG_SOFT)
                    icon_label2.image = None
                    icon_more_label.configure(text="", bg=BG_SOFT)
                    info_label.configure(text="", bg=BG_SOFT)
                    amount_label.configure(text="", bg=BG_SOFT)
                    self.calendar_cell_meta[(r, c)] = {"day": 0, "events": []}
                    continue

                events = day_map.get(day_num, [])
                self.calendar_cell_meta[(r, c)] = {"day": day_num, "events": events}
                if events:
                    preview = self._format_calendar_event_text(events)
                    day_total = sum(s.price for s in events)
                    icon_images = []
                    for ev in events:
                        if ev.icon_path:
                            img = self._load_icon_image(ev.icon_path)
                            if img is not None:
                                icon_images.append(img)
                        if len(icon_images) >= 2:
                            break
                    cell.configure(bg=ACCENT_SOFT)
                    day_label.configure(text=str(day_num), bg=ACCENT_SOFT, fg=ACCENT_TEXT)
                    if len(icon_images) >= 1:
                        icon_label.configure(image=icon_images[0], text="", bg=ACCENT_SOFT)
                        icon_label.image = icon_images[0]
                    else:
                        icon_label.configure(image="", text="", bg=ACCENT_SOFT)
                        icon_label.image = None
                    if len(icon_images) >= 2:
                        icon_label2.configure(image=icon_images[1], text="", bg=ACCENT_SOFT)
                        icon_label2.image = icon_images[1]
                    else:
                        icon_label2.configure(image="", text="", bg=ACCENT_SOFT)
                        icon_label2.image = None
                    icon_more_label.configure(
                        text=(f"+{len(events) - 2}" if len(events) > 2 else ""),
                        bg=ACCENT_SOFT,
                    )
                    wrap = max(96, cell.winfo_width() - 12)
                    info_label.configure(text=preview, bg=ACCENT_SOFT, wraplength=wrap)
                    amount_label.configure(text=f"¥{day_total:.2f}", bg=ACCENT_SOFT)
                else:
                    cell.configure(bg=BG_CARD)
                    day_label.configure(text=str(day_num), bg=BG_CARD, fg=TEXT_SUB)
                    icon_label.configure(image="", text="", bg=BG_CARD)
                    icon_label.image = None
                    icon_label2.configure(image="", text="", bg=BG_CARD)
                    icon_label2.image = None
                    icon_more_label.configure(text="", bg=BG_CARD)
                    info_label.configure(text="", bg=BG_CARD)
                    amount_label.configure(text="", bg=BG_CARD)

    def _on_calendar_cell_enter(self, event, row: int, col: int) -> None:
        self._cancel_calendar_tooltip_after()
        self.calendar_tooltip_after = self.after(180, lambda: self._show_calendar_tooltip(event, row, col))

    def _on_calendar_cell_motion(self, event, row: int, col: int) -> None:
        if self.calendar_tooltip and self.calendar_tooltip.winfo_exists():
            x = event.x_root + 14
            y = event.y_root + 14
            self.calendar_tooltip.geometry(f"+{x}+{y}")

    def _on_calendar_cell_leave(self, _event=None) -> None:
        self._cancel_calendar_tooltip_after()
        self._hide_calendar_tooltip()

    def _cancel_calendar_tooltip_after(self) -> None:
        if self.calendar_tooltip_after is not None:
            try:
                self.after_cancel(self.calendar_tooltip_after)
            except Exception:
                pass
            self.calendar_tooltip_after = None

    def _hide_calendar_tooltip(self) -> None:
        if self.calendar_tooltip is not None and self.calendar_tooltip.winfo_exists():
            self.calendar_tooltip.destroy()
        self.calendar_tooltip = None

    def _show_calendar_tooltip(self, event, row: int, col: int) -> None:
        self.calendar_tooltip_after = None
        meta = self.calendar_cell_meta.get((row, col), {"day": 0, "events": []})
        events = meta.get("events") or []
        day_num = meta.get("day", 0)
        if not events or not day_num:
            return

        self._hide_calendar_tooltip()
        tip = tk.Toplevel(self)
        tip.overrideredirect(True)
        tip.attributes("-topmost", True)
        tip.configure(bg=BORDER)

        body = tk.Frame(tip, bg=BG_CARD, padx=10, pady=8)
        body.pack(fill="both", expand=True, padx=1, pady=1)
        tk.Label(
            body,
            text=f"{self.calendar_year}-{self.calendar_month:02d}-{day_num:02d}  共{len(events)}项",
            bg=BG_CARD,
            fg=TEXT_SUB,
            font=("Arial", 11, "bold"),
            anchor="w",
        ).pack(anchor="w")
        for s in events:
            tk.Label(
                body,
                text=f"• {s.service}   ¥{s.price:.2f}",
                bg=BG_CARD,
                fg=TEXT_MAIN,
                font=("Arial", 11),
                anchor="w",
            ).pack(anchor="w", pady=(2, 0))

        total = sum(s.price for s in events)
        tk.Label(
            body,
            text=f"合计：¥{total:.2f}",
            bg=BG_CARD,
            fg="#166534",
            font=("Arial", 11, "bold"),
            anchor="w",
        ).pack(anchor="w", pady=(6, 0))

        x = event.x_root + 14
        y = event.y_root + 14
        tip.geometry(f"+{x}+{y}")
        self.calendar_tooltip = tip

    def _build_reminder_tab(self) -> None:
        self.reminder_tree = ttk.Treeview(
            self.reminder_frame,
            columns=("service", "renew", "days", "level", "status", "remark"),
            show="headings",
            height=24,
            style="Glass.Treeview",
        )
        for col, title, w in (
            ("service", "服务", 180),
            ("renew", "下次续费日期", 150),
            ("days", "剩余天数", 100),
            ("level", "提醒级别", 120),
            ("status", "状态", 120),
            ("remark", "备注", 280),
        ):
            self.reminder_tree.heading(col, text=title)
            self.reminder_tree.column(col, width=w, anchor="center")
        self.reminder_tree.pack(fill="both", expand=True, padx=10, pady=10)

    def add_item(self) -> None:
        dialog = SubscriptionDialog(self, "新增订阅")
        self.wait_window(dialog)
        if dialog.result:
            self._push_undo_snapshot()
            self.store.subscriptions.append(dialog.result)
            self.refresh_all()

    def _get_selected_data_index(self) -> Optional[int]:
        if hasattr(self, "draw_table"):
            idx = self.draw_table.get_selected_source_index()
            if idx is not None:
                return int(idx)
        if hasattr(self, "tree"):
            sel = self.tree.selection()
            if sel:
                return int(sel[0])
        return None

    def edit_item(self) -> None:
        index = self._get_selected_data_index()
        if index is None:
            messagebox.showwarning("提示", "请先选择一条订阅")
            return
        dialog = SubscriptionDialog(self, "编辑订阅", self.store.subscriptions[index])
        self.wait_window(dialog)
        if dialog.result:
            self._push_undo_snapshot()
            self.store.subscriptions[index] = dialog.result
            self.refresh_all()

    def delete_item(self) -> None:
        index = self._get_selected_data_index()
        if index is None:
            messagebox.showwarning("提示", "请先选择一条订阅")
            return
        ok = messagebox.askyesno("确认删除", f"确认删除：{self.store.subscriptions[index].service} ?")
        if ok:
            self._push_undo_snapshot()
            self.store.subscriptions.pop(index)
            self.refresh_all()

    def import_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="选择订阅Excel",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        try:
            self._push_undo_snapshot()
            self.store.import_from_excel(Path(path))
            self._sync_dynamic_options_from_data()
            self.refresh_all()
            messagebox.showinfo("导入成功", "已从 Excel 导入订阅数据")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def export_excel(self) -> None:
        path = filedialog.asksaveasfilename(
            title="导出订阅Excel",
            defaultextension=".xlsx",
            initialfile="订阅管理系统_导出.xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        try:
            self.store.export_to_excel(Path(path))
            messagebox.showinfo("导出成功", f"文件已导出到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def save(self) -> None:
        self.store.save()
        messagebox.showinfo("保存成功", f"数据已保存到:\n{DATA_FILE}")

    def _sync_dynamic_options_from_data(self) -> None:
        for s in self.store.subscriptions:
            if s.category and s.category not in self.dynamic_category_options:
                self.dynamic_category_options.append(s.category)
            if s.cycle and s.cycle not in self.dynamic_cycle_options:
                self.dynamic_cycle_options.append(s.cycle)

    def _push_undo_snapshot(self) -> None:
        snapshot = [x.to_dict() for x in self.store.subscriptions]
        self.undo_stack.append(snapshot)
        if len(self.undo_stack) > 30:
            self.undo_stack.pop(0)
        self._refresh_undo_button()

    def _refresh_undo_button(self) -> None:
        if hasattr(self, "undo_btn"):
            self.undo_btn.configure(state=("normal" if self.undo_stack else "disabled"))

    def undo_last(self) -> None:
        if not self.undo_stack:
            return
        snapshot = self.undo_stack.pop()
        self.store.subscriptions = [Subscription.from_dict(x) for x in snapshot]
        self.refresh_all()
        self._refresh_undo_button()

    def download_template(self) -> None:
        path = filedialog.asksaveasfilename(
            title="下载订阅模板",
            defaultextension=".xlsx",
            initialfile="订阅管理模板.xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "订阅清单"
        ws.append(
            [
                "服务",
                "类别",
                "计费周期",
                "价格(元)",
                "下次续费日期",
                "是否常用",
                "重要性",
                "状态",
                "备注",
                "图标路径",
            ]
        )
        ws.append(["示例：ChatGPT", "AI", "月", 20, "2026-04-16", "Y", "高", "活跃", "", ""])
        wb.save(path)
        messagebox.showinfo("下载完成", f"模板已保存到:\n{path}")

    def reset_filters(self) -> None:
        self.keyword_var.set("")
        self.column_filters.clear()
        self.filter_hint_var.set("提示：点击表头可筛选（类别/周期/常用/重要性/状态/续费提醒）")
        self.refresh_list()

    def _update_category_filter_options(self) -> None:
        # Column-header filtering does not need UI combobox refresh.
        return

    def _on_tree_click_header(self, event) -> None:
        region = self.tree.identify("region", event.x, event.y)
        if region != "heading":
            self._hide_header_filter()
            return
        col_id = self.tree.identify_column(event.x)  # like #1, #2 ...
        if not col_id:
            return
        col_index = int(col_id.replace("#", "")) - 1
        columns = list(self.tree["columns"])
        if col_index < 0 or col_index >= len(columns):
            return
        col_name = columns[col_index]
        if col_name not in self.filterable_columns:
            return
        self._show_header_dropdown(col_name, event.x)
        return "break"

    def _show_header_dropdown(self, column: str, x: int) -> None:
        self._hide_header_filter()
        options = self._get_header_filter_options(column)
        current = self.column_filters.get(column, "全部")
        self.header_filter_panel = tk.Frame(
            self.list_table_wrap,
            bg=BG_CARD,
            highlightbackground=BORDER,
            highlightthickness=1,
            bd=0,
        )
        self.header_filter_list = tk.Listbox(
            self.header_filter_panel,
            exportselection=False,
            height=min(8, max(3, len(options))),
            bg=BG_CARD,
            fg=TEXT_MAIN,
            highlightthickness=0,
            bd=0,
            selectbackground=ACCENT_SOFT,
            selectforeground=TEXT_MAIN,
        )
        for i, opt in enumerate(options):
            self.header_filter_list.insert("end", opt)
            if opt == current:
                self.header_filter_list.selection_set(i)
        if not self.header_filter_list.curselection():
            self.header_filter_list.selection_set(0)
        self.header_filter_list.pack(fill="both", expand=True, padx=2, pady=2)
        place_x = max(0, min(x - 30, max(0, self.list_table_wrap.winfo_width() - 180)))
        self.header_filter_panel.place(x=place_x, y=28, width=180)
        self.header_filter_list.focus_set()
        self.header_filter_list.bind(
            "<<ListboxSelect>>",
            lambda _e, c=column: self._apply_header_filter(c, self.header_filter_list.get(self.header_filter_list.curselection()[0]))
            if self.header_filter_list.curselection()
            else None,
        )
        self.header_filter_list.bind("<Escape>", lambda _e: self._hide_header_filter())
        self.header_filter_list.bind("<FocusOut>", lambda _e: self._hide_header_filter())
        self.after(0, self._bind_header_filter_outside_close)

    def _hide_header_filter(self) -> None:
        if hasattr(self, "header_filter_panel") and self.header_filter_panel.winfo_exists():
            self.header_filter_panel.destroy()
        if self.header_filter_close_bindid:
            self.unbind("<Button-1>", self.header_filter_close_bindid)
            self.header_filter_close_bindid = None

    def _bind_header_filter_outside_close(self) -> None:
        if self.header_filter_close_bindid:
            self.unbind("<Button-1>", self.header_filter_close_bindid)
            self.header_filter_close_bindid = None
        self.header_filter_close_bindid = self.bind("<Button-1>", self._on_global_click_close_header_filter, add="+")

    def _on_global_click_close_header_filter(self, event) -> None:
        if not hasattr(self, "header_filter_panel") or not self.header_filter_panel.winfo_exists():
            return
        panel = self.header_filter_panel
        w = event.widget
        while w is not None:
            if w == panel:
                return
            w = w.master
        self._hide_header_filter()

    def _get_header_filter_options(self, column: str) -> List[str]:
        def v(s: Subscription) -> str:
            if column == "category":
                return s.category
            if column == "cycle":
                return s.cycle
            if column == "frequent":
                return s.frequent
            if column == "importance":
                return s.importance
            if column == "status":
                return s.status
            if column == "reminder":
                return s.reminder
            return ""

        values = sorted({v(s) for s in self.store.subscriptions if v(s)})
        return ["全部"] + values

    def _apply_header_filter(self, column: str, selected: str) -> None:
        if selected == "全部":
            self.column_filters.pop(column, None)
        else:
            self.column_filters[column] = selected
        if self.column_filters:
            summary = " | ".join(f"{k}:{v}" for k, v in self.column_filters.items())
            self.filter_hint_var.set(f"已筛选：{summary}")
        else:
            self.filter_hint_var.set("提示：点击表头可筛选（类别/周期/常用/重要性/状态/续费提醒）")
        self._hide_header_filter()
        self.refresh_list()

    def _on_column_header_click(self, column: str) -> None:
        # Kept only for backward compatibility; filtering now uses inline dropdown on header click.
        self._show_header_dropdown(column, 40)

    def _matches_filters(self, s: Subscription) -> bool:
        keyword = self.keyword_var.get().strip().lower()
        if keyword:
            haystack = f"{s.service} {s.category} {s.remark}".lower()
            if keyword not in haystack:
                return False
        for col, expected in self.column_filters.items():
            if col == "category" and s.category != expected:
                return False
            if col == "cycle" and s.cycle != expected:
                return False
            if col == "frequent" and s.frequent != expected:
                return False
            if col == "importance" and s.importance != expected:
                return False
            if col == "status" and s.status != expected:
                return False
            if col == "reminder" and s.reminder != expected:
                return False
        return True

    def _load_icon_image(self, icon_path: str) -> Optional[tk.PhotoImage]:
        if not icon_path:
            return None
        if icon_path in self.icon_cache:
            return self.icon_cache[icon_path]
        p = Path(icon_path)
        if not p.exists():
            return None
        try:
            img = tk.PhotoImage(file=str(p))
            max_side = max(img.width(), img.height())
            if max_side > 18:
                scale = max(1, max_side // 18)
                img = img.subsample(scale, scale)
            self.icon_cache[icon_path] = img
            return img
        except tk.TclError:
            if Image is None or ImageTk is None:
                return None
            try:
                pil_img = Image.open(str(p)).convert("RGBA")
                pil_img.thumbnail((18, 18))
                img2 = ImageTk.PhotoImage(pil_img)
                self.icon_cache[icon_path] = img2
                return img2
            except Exception:
                return None

    def _auto_fit_list_columns(self) -> None:
        try:
            body_font = tkfont.nametofont("TkDefaultFont")
            heading_font = tkfont.nametofont("TkHeadingFont")
        except Exception:
            body_font = tkfont.Font(family="Arial", size=12)
            heading_font = tkfont.Font(family="Arial", size=12, weight="bold")

        pad = 24
        max_width = 260
        min_map = {
            "seq": 42,
            "service": 96,
            "category": 72,
            "cycle": 62,
            "price": 72,
            "monthly": 72,
            "yearly": 82,
            "renew": 110,
            "frequent": 62,
            "importance": 72,
            "status": 72,
            "saving": 96,
            "reminder": 110,
            "remark": 90,
        }

        for col in self.tree["columns"]:
            head_text = self.tree.heading(col, "text")
            width = heading_font.measure(str(head_text)) + pad
            for item in self.tree.get_children():
                raw = self.tree.set(item, col)
                width = max(width, body_font.measure(str(raw)) + pad)
            width = max(min_map.get(col, 60), min(max_width, width))
            self.tree.column(col, width=width)
        self.tree.column("#0", width=56)

    def refresh_all(self) -> None:
        self._sync_dynamic_options_from_data()
        self._update_category_filter_options()
        self.refresh_list()
        self.refresh_dashboard()
        self.refresh_reminders()
        self._refresh_undo_button()

    def refresh_list(self) -> None:
        shown: List[Subscription] = []
        draw_rows = []
        display_idx = 0
        for idx, s in enumerate(self.store.subscriptions):
            if not self._matches_filters(s):
                continue
            display_idx += 1
            icon = self._load_icon_image(s.icon_path)
            draw_rows.append(
                {
                    "__index": idx,
                    "__icon": icon,
                    "seq": str(display_idx),
                    "icon": "",
                    "service": s.service,
                    "category": s.category,
                    "cycle": s.cycle,
                    "price": f"{s.price:.2f}",
                    "monthly": f"{s.monthly_cost:.2f}",
                    "yearly": f"{s.yearly_cost:.2f}",
                    "renew": s.renew_date.isoformat() if s.renew_date else "",
                    "frequent": s.frequent,
                    "importance": s.importance,
                    "status": s.status,
                    "saving": f"{s.cancel_saving:.2f}",
                    "reminder": s.reminder,
                    "remark": s.remark,
                }
            )
            shown.append(s)
        if hasattr(self, "draw_table"):
            self.draw_table.set_rows(draw_rows, self.icon_cache)

        self.list_summary_vars["count"].set(str(len(shown)))
        self.list_summary_vars["month"].set(f"{sum(s.monthly_cost for s in shown):.2f}")
        self.list_summary_vars["year"].set(f"{sum(s.yearly_cost for s in shown):.2f}")
        self.list_summary_vars["saving"].set(f"{sum(s.cancel_saving for s in shown):.2f}")

    def refresh_dashboard(self) -> None:
        subs = self.store.subscriptions
        total_month = sum(s.monthly_cost for s in subs)
        total_year = sum(s.yearly_cost for s in subs)
        total_saving = sum(s.cancel_saving for s in subs)

        self.metrics["month"].set(f"¥ {total_month:.2f}")
        self.metrics["year"].set(f"¥ {total_year:.2f}")
        self.metrics["saving"].set(f"¥ {total_saving:.2f}")
        self.metrics["active"].set(str(sum(1 for s in subs if s.status == "活跃")))
        self.metrics["pending"].set(str(sum(1 for s in subs if s.status == "待取消")))
        self.metrics["low_use"].set(str(sum(1 for s in subs if s.frequent == "N")))

        for item in self.category_tree.get_children():
            self.category_tree.delete(item)
        category_map: Dict[str, List[float]] = {}
        for s in subs:
            bucket = category_map.setdefault(s.category or "未分类", [0.0, 0.0])
            bucket[0] += s.monthly_cost
            bucket[1] += s.yearly_cost
        for category, values in sorted(category_map.items(), key=lambda kv: kv[1][1], reverse=True):
            self.category_tree.insert("", "end", values=(category, f"{values[0]:.2f}", f"{values[1]:.2f}"))

        for item in self.top_tree.get_children():
            self.top_tree.delete(item)
        top5 = sorted(subs, key=lambda s: s.yearly_cost, reverse=True)[:5]
        for i, s in enumerate(top5, start=1):
            self.top_tree.insert("", "end", values=(i, s.service, f"{s.yearly_cost:.2f}"))
        self._refresh_dashboard_calendar()

    def refresh_reminders(self) -> None:
        for item in self.reminder_tree.get_children():
            self.reminder_tree.delete(item)

        def sort_key(x: Subscription):
            d = x.days_left
            return (999999 if d is None else d)

        for s in sorted(self.store.subscriptions, key=sort_key):
            days = "" if s.days_left is None else str(s.days_left)
            self.reminder_tree.insert(
                "",
                "end",
                values=(
                    s.service,
                    s.renew_date.isoformat() if s.renew_date else "",
                    days,
                    s.reminder_level,
                    s.status,
                    s.remark,
                ),
                tags=(s.reminder_level,),
            )

        self.reminder_tree.tag_configure("紧急", foreground="#c62828")
        self.reminder_tree.tag_configure("关注", foreground="#ef6c00")
        self.reminder_tree.tag_configure("已过期", foreground="#6a1b9a")


def main() -> None:
    install_crash_handler()
    store = DataStore()
    store.load()

    app = App(store)
    app.mainloop()


if __name__ == "__main__":
    main()
